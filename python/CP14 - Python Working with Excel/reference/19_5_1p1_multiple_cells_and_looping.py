# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

'''
Practice:

Import the mock_grades file.

PART 1:
    Change the grade of anyone with a "C-" to be an "F" on the active sheet.
    Save the results as a new workbook called "mock_grades_altered"

PART 2:
    If you can successfully do that, then alter your code to change "C-" to "F" on all of the sheets
    in mock_grades. Save the results as a new workbook called "mock_grades_altered"

'''

import openpyxl

# load in mock_grades again
external_workbook = openpyxl.load_workbook("mock_grades.xlsx")

# PART 1:
# optionally store the active worksheet in a new variable
current_sheet = external_workbook.active

for row in current_sheet.iter_rows():
    for cell in row:
        if cell.value == "C-":
            cell.value = "F"

external_workbook.save(filename="mock_grades_altered.xlsx")      


# PART 2:

for worksheet in external_workbook.worksheets:
    for row in worksheet.iter_rows():
        if row[1].value == "C-":
            row[1].value = "F"

external_workbook.save(filename="mock_grades_altered.xlsx")      
external_workbook.close()