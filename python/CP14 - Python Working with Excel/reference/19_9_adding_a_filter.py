# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

import openpyxl

externalWorkbook = openpyxl.load_workbook("Mock_Grades.xlsx")

# on the worksheet, just add .auto_filter.ref = "coordinate range"
# in Mock_Grades, the data goes from A1:B51
externalWorkbook.active.auto_filter.ref = 'A1:B51'

# what if you don't know how many rows? (This is useful for your project)
# use max_row:
furthestRowWithData = externalWorkbook.active.max_row
print(f"furthest row with data: {furthestRowWithData}")
externalWorkbook.active.auto_filter.ref = f'A1:B{furthestRowWithData}'


# What if you don't know how many columns? (Note, on your project, you know exactly how many columns, so you don't need this, but you could use it)
# You can use max_column:

furthestColumnWithData = externalWorkbook.active.max_column
print(f"furthest column with data: {furthestRowWithData}")

# notice how it is a number? We need a Letter instead of a number, so you could use this:
from openpyxl.utils import get_column_letter
column_letter = get_column_letter(furthestColumnWithData)

#now you can combine them to get this:
externalWorkbook.active.auto_filter.ref = f'A1:{column_letter}{furthestRowWithData}'

externalWorkbook.save(filename="Mock_Grades_Filter.xlsx")
externalWorkbook.close()