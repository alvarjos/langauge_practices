# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

import openpyxl

# load in mock_grades again
external_workbook = openpyxl.load_workbook("mock_grades.xlsx")

# optionally store the active worksheet in a new variable
current_sheet = external_workbook.active


# use .iter_rows() in a for loop to get all the rows. print out the row object (not useful yet)
for row in current_sheet.iter_rows():
    print(row)


# use values_only = True to only return the values of the row. print out the whole row
for row in current_sheet.iter_rows(values_only = True):
    print(row)

# now print out each individual cell. Use values_only = true. Remember in each row, there are multiple cells
for row in current_sheet.iter_rows(values_only = True):
    for cell in row:
        print(cell)


# Do the same thing, but now do not use values_only = True.
# For every cell, print out the coordinate (use .coordinate on the cell) and the value (use .value)
for row in current_sheet.iter_rows():
    for cell in row:
        print(f"this is the coordinate: {cell.coordinate}, and this is the value: {cell.value}")


# you can use min_row, min_col, max_col, max_row to only go through a section of the worksheet
# only get the first 10 rows of the 2nd and 3rd columns
for row in current_sheet.iter_rows(min_row = 1, max_row = 10, min_col = 2, max_col = 3) :
    for obj_cell in row :
        print(obj_cell.value)


# Could also get the specific cells you want with a coordinate range like curentSheet["A1:B51"] and loop through that.
range = current_sheet["A1:B51"]

for row in range:
    for cell in row:
        print(cell.value)
