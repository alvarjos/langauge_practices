# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

# importing openpyxl import Workbook
import openpyxl

#Make a new Workbook by calling the Workbook() constructor. Workbook is a class inside the openpyxl library
my_workbook = openpyxl.Workbook()

# you can use .title to get the title of a sheet. What is the title of the active sheet?
# remember my_workbook.active
print(my_workbook.active.title)

# change the title of the default sheet to "Students"
my_workbook.active.title = "Students"

# create another sheet called "Classes" using .create_sheet("Sheet name")
my_workbook.create_sheet("Classes")

# note: you can also remove a sheet using .remove, but I won't do that now.

#change the active sheet to classes. Access a specific sheet using my_workbook["Sheet Name"]
my_workbook.active = my_workbook["Classes"]

# Add something to the first row, like Class Name and then Department.
# Use square brackets to access a position in the sheet, like ["A1"]
my_workbook.active["A1"] = "Class Name"
my_workbook.active["B1"] = "Department"

# add stuff to the second row
my_workbook.active["A2"] = "IS 303"
my_workbook.active["B2"] = "Information Systems"

# put the cell A2 in a variable called cell
cell = my_workbook.active["A2"]

# for any cell, you can access its value, row, column, and coordinate (column and row)
# print out each of those (use .value to get the value, etc.)
print(cell.value)
print(cell.row)
print(cell.column) # Notice that it gives a number instead of a letter
print(cell.coordinate)

# you can use the .cell(row_number, column number) if you don't know the exact letter of the column you want
# unlike Pandas, this starts with 1 instead of 0.
# try printing the value of the 1st row, 2nd column.
print(my_workbook.active.cell(1,2).value) 

# note you can also access sheets that aren't the active sheet
# by using my_workbook["Sheet Name"]["Coordinate (like A1, etc.)"]
# Try changing the "Students" sheet so that the A1 value says "Name"
my_workbook["Students"]["A1"] = "Name"


# other useful things, on a worksheet, you can use .max_row to get the furthest row with data, same with .max_col
print(my_workbook.active.max_row)
print(my_workbook.active.max_column)

# use .save(filename = "filename.xlsx") to actually create a file
# create an excel file named example2
my_workbook.save(filename="example2.xlsx") 

my_workbook.close()