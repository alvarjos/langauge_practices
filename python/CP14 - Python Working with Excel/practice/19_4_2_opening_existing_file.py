# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

# importing openpyxl import Workbook
import openpyxl

# you can open up an existing workbook by using openpyxl.load_workbook
# in the load_workbook function, you need to provide a file path to the excel spreadsheet
# I've given you an excel workbook called mock_grades. You can right click the file in the explorer
# window to the left to get the file path. Relative file paths are nice for it working on anyone's computer, but
# it depends on where you are running this python file from. Absolute paths will only work on your computer, but
# it doesn't matter where you are running the python file from.

# Absolute Path 
# external_wb = openpyxl.load_workbook("C:\Users\jange\OneDrive\BYU\WINTER 2024\IS 303 Notes\CP14 - Python Working with Excel\mock_grades.xlsx")

# Relative Path
external_wb = openpyxl.load_workbook("CP14 - Python Working with Excel\mock_grades.xlsx")

# use .sheetnames to see the names of the sheets of the whole workbook
print(external_wb.sheetnames)

# that might come in handy when trying to do something to multiple tables.
# loop through each of the sheet names and print out the intidivudal sheet name
for sheets in external_wb.sheetnames:
    print(sheets)
    

# if you want the entire worksheet object, not just the name of the worksheet,
# you can use .worksheets
print(external_wb.worksheets)

# try printing out the value of A2 from each worksheet:
for worksheet in external_wb.worksheets:
    print(worksheet["A2"].value)

# .append
# use it to add a new row to the worksheet on the next empty row.
# try appending a list that has "New test student" and "A++"' as data


# save the file as mock_grades_altered and close the Workbook you opened.

list_to_append = ["Test Student", "A++"]
external_wb["IS 201"].append(list_to_append)


external_wb.save("mockGradesAltered.xlsx")
external_wb.close()